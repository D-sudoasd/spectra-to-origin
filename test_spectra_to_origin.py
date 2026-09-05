"""Tests import shipped functions from spectra_to_origin.py."""

from __future__ import annotations

import contextlib
import ctypes
import inspect
import io
import os
import shutil
import subprocess
import sys
import tempfile
import tkinter as tk
import unittest
from pathlib import Path
from unittest import mock

import spectra_to_origin as sto

TI2448 = Path(r"D:\Backup\桌面\Aging_data_analysis\Ti2448_时效eta_83keV\01_1D谱线_2theta")

TI2448_STEMS = [
    "01_400C_00.50h_eta0.61000",
    "02_400C_01.00h_eta0.64000",
    "03_400C_02.00h_eta0.68419",
    "04_400C_04.00h_eta0.73000",
    "05_400C_12.00h_eta0.77200",
    "06_400C_24.00h_eta0.78634",
    "07_400C_48.00h_eta0.79925",
    "08_400C_96.00h_eta0.83700",
    "09_450C_00.50h_eta0.72828",
    "10_450C_01.00h_eta0.77344",
    "11_450C_02.00h_eta0.82183",
    "12_450C_04.00h_eta0.85086",
    "13_450C_12.00h_eta0.90140",
    "14_450C_24.00h_eta0.91000",
    "15_450C_48.00h_eta0.91645",
    "16_500C_00.25h_eta0.85000",
    "17_500C_00.50h_eta0.88174",
    "18_500C_01.00h_eta0.90677",
    "19_500C_02.00h_eta0.91430",
    "20_500C_04.00h_eta0.92500",
    "21_500C_12.00h_eta0.92935",
    "22_500C_24.00h_eta0.93200",
    "23_500C_48.00h_eta0.93400",
]


def _spectrum(name: str, x: tuple[str, ...], y: tuple[str, ...]) -> sto.Spectrum:
    return sto.Spectrum(
        path=Path(f"{name}.txt"),
        x_text=x,
        y_text=y,
        long_name=name,
        comment="",
        temperature_tag=None,
    )


def _write_xy(path: Path, pairs: list[tuple[str, str]]) -> None:
    lines = [f"{x}\t{y}" for x, y in pairs]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


class LayoutTests(unittest.TestCase):
    def test_identical_x_infers_xyyy(self) -> None:
        x = ("0.0", "1.0", "2.0")
        spectra = [
            _spectrum("a", x, ("1", "2", "3")),
            _spectrum("b", x, ("4", "5", "6")),
        ]
        self.assertTrue(sto.shared_x_grid(spectra))
        self.assertEqual(sto.infer_layout(spectra), "XYYY")
        self.assertEqual(sto.resolve_layout(spectra, "auto"), "XYYY")
        table = sto.build_table("all", spectra, "auto")
        self.assertEqual(table.long_names[0], "2theta")
        self.assertEqual(len(table.long_names), 3)

    def test_different_x_infers_xyxy(self) -> None:
        spectra = [
            _spectrum("a", ("0", "1", "2"), ("1", "2", "3")),
            _spectrum("b", ("10", "11", "12"), ("4", "5", "6")),
        ]
        self.assertFalse(sto.shared_x_grid(spectra))
        self.assertEqual(sto.infer_layout(spectra), "XYXY")
        self.assertEqual(sto.resolve_layout(spectra, "auto"), "XYXY")
        table = sto.build_table("all", spectra, "auto")
        self.assertEqual(len(table.long_names), 4)
        self.assertEqual(table.rows[0][0], "0")
        self.assertEqual(table.rows[0][2], "10")

    def test_xyyy_build_raises_on_mismatch(self) -> None:
        spectra = [
            _spectrum("a", ("0", "1", "2"), ("1", "2", "3")),
            _spectrum("b", ("10", "11", "12"), ("4", "5", "6")),
        ]
        with self.assertRaisesRegex(ValueError, "X 网格不一致"):
            sto.build_table("all", spectra, "XYYY")


class GroupTests(unittest.TestCase):
    def test_temperature_stems_are_not_dumped_into_other(self) -> None:
        groups = sto.suggest_filename_groups(TI2448_STEMS)
        names = [name for name, _indexes in groups]
        indexes = [idx for _name, idxs in groups for idx in idxs]
        self.assertEqual(sorted(indexes), list(range(len(TI2448_STEMS))))
        self.assertNotIn("other", names)
        self.assertGreaterEqual(len(groups), 2)
        self.assertEqual(set(names), {"400C", "450C", "500C"})
        sizes = {name: len(idxs) for name, idxs in groups}
        self.assertEqual(sizes["400C"], 8)
        self.assertEqual(sizes["450C"], 7)
        self.assertEqual(sizes["500C"], 8)

    def test_non_temperature_token_family(self) -> None:
        stems = [
            "alloy_alpha_01",
            "alloy_alpha_02",
            "alloy_beta_01",
            "alloy_beta_02",
            "alloy_gamma_01",
            "alloy_gamma_02",
        ]
        groups = sto.suggest_filename_groups(stems)
        names = [name for name, _indexes in groups]
        indexes = [idx for _name, idxs in groups for idx in idxs]
        self.assertEqual(sorted(indexes), list(range(6)))
        self.assertNotIn("other", names)
        self.assertEqual(set(names), {"alpha", "beta", "gamma"})
        for _name, idxs in groups:
            self.assertEqual(len(idxs), 2)

    def test_manual_partition_3_and_4_feed_sheet_specs(self) -> None:
        x = ("0", "1")
        spectra = [_spectrum(f"s{i:02d}", x, ("1", "2")) for i in range(10)]
        groups3 = sto.partition_spectra(spectra, 3)
        specs3 = sto.build_sheet_specs(groups3, "XYYY")
        self.assertEqual(len(specs3), 3)
        self.assertEqual([spec.name for spec in specs3], ["group1", "group2", "group3"])
        flat3 = [item.path for spec in specs3 for item in spec.spectra]
        self.assertEqual(flat3, [item.path for item in spectra])
        self.assertEqual(sum(len(spec.spectra) for spec in specs3), 10)

        groups4 = sto.partition_spectra(spectra, 4)
        specs4 = sto.build_sheet_specs(groups4, "auto")
        self.assertEqual(len(specs4), 4)
        self.assertEqual([spec.name for spec in specs4], ["group1", "group2", "group3", "group4"])
        flat4 = [item.path for spec in specs4 for item in spec.spectra]
        self.assertEqual(flat4, [item.path for item in spectra])
        self.assertTrue(all(spec.layout == "XYYY" for spec in specs4))

    def test_unique_origin_names_do_not_collide(self) -> None:
        names = sto.unique_origin_names(["400C", "400C", "400 C extra", "group1", "group1"])
        lowered = [name.lower() for name in names]
        self.assertEqual(len(lowered), len(set(lowered)))
        self.assertTrue(all(len(name) <= 13 for name in names))
        self.assertTrue(all(name[0].isalpha() for name in names))

    def test_commit_exported_file_cross_drive(self) -> None:
        c_root = Path(tempfile.gettempdir()).resolve()
        d_root = Path(sto.__file__).resolve().parent / "_test_out"
        d_root.mkdir(parents=True, exist_ok=True)
        src_dir = Path(tempfile.mkdtemp(prefix="opju_src_", dir=str(c_root)))
        src = src_dir / "export.opju"
        payload = b"OPJU-CROSS-DRIVE-TEST"
        src.write_bytes(payload)
        dest = d_root / "committed.opju"
        if dest.exists():
            dest.unlink()
        self.assertNotEqual(src.drive.lower(), dest.drive.lower())
        result = sto.commit_exported_file(src, dest)
        self.assertEqual(result, dest.resolve())
        self.assertTrue(dest.exists())
        self.assertEqual(dest.read_bytes(), payload)
        self.assertFalse(src.exists())
        dest.unlink()
        shutil.rmtree(src_dir, ignore_errors=True)


class DropLoadTests(unittest.TestCase):
    def test_simulated_drop_dedupes_temp_copies(self) -> None:
        self.assertTrue(TI2448.is_dir(), f"missing sample folder {TI2448}")
        payload = [
            str(TI2448),
            str(TI2448 / "按温度"),
            str(TI2448 / "01_400C_00.50h_eta0.61000.txt"),
        ]
        loaded = sto.load_from_drop_payload([], payload)
        names = [path.name.lower() for path in loaded]
        self.assertEqual(len(loaded), 24)
        self.assertEqual(len(set(names)), 24)
        self.assertIn("00_hr_00.00h_eta0.00000.txt", names)
        self.assertFalse(any(sto.TEMP_COPY_DIR in path.parts for path in loaded))

    def test_newline_and_quoted_drop_payload(self) -> None:
        folder = TI2448
        text = f'"{folder}"\r\n{folder / "按温度"}'
        paths = sto.parse_drop_paths(text)
        self.assertEqual(len(paths), 2)
        loaded = sto.load_from_drop_payload([], text)
        self.assertEqual(len(loaded), 24)
        self.assertIn("00_HR_00.00h_eta0.00000.txt", [path.name for path in loaded])

    def test_window_drop_is_wired_to_load_path(self) -> None:
        source = inspect.getsource(sto.SpectraToOriginApp)
        self.assertIn("enable_windows_file_drop", source)
        self.assertIn("_on_drop_paths", source)
        self.assertIn("load_from_drop_payload", inspect.getsource(sto.SpectraToOriginApp._add_paths))
        self.assertIn("WM_DROPFILES", inspect.getsource(sto._enable_windows_file_drop))
        self.assertIn("_enable_windows_file_drop", inspect.getsource(sto.enable_windows_file_drop))
        add_source = inspect.getsource(sto.SpectraToOriginApp._build)
        self.assertIn("添加文件", add_source)
        self.assertIn("添加文件夹", add_source)
        self.assertIn("按文件名分组", add_source)
        self.assertIn("应用均分", add_source)
        self.assertIn("自动", add_source)

    def test_enable_drop_on_real_tk_update_does_not_overflow(self) -> None:
        ignored: list[BaseException] = []

        def _unraisable(unraisable) -> None:
            ignored.append(unraisable.exc_value)

        root = tk.Tk()
        root.withdraw()
        previous = sys.unraisablehook
        sys.unraisablehook = _unraisable
        try:
            hooked = sto.enable_windows_file_drop(root, lambda _paths: None)
            self.assertTrue(hooked)
            self.assertTrue(callable(getattr(root, "_drop_wndproc", None)))
            try:
                for _ in range(8):
                    root.update_idletasks()
                    root.update()
            except (ctypes.ArgumentError, OverflowError) as exc:
                self.fail(f"drop hook ctypes error during Tk update: {exc}")
            ctypes_errors = [
                exc
                for exc in ignored
                if isinstance(exc, (ctypes.ArgumentError, OverflowError))
            ]
            self.assertEqual(ctypes_errors, [], ctypes_errors)
        finally:
            sys.unraisablehook = previous
            root.destroy()


class FixtureRoundTripTests(unittest.TestCase):
    def test_xlsx_only_same_x_and_diff_x(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            root = Path(raw)
            same = root / "same"
            diff = root / "diff"
            same.mkdir()
            diff.mkdir()
            _write_xy(same / "a.txt", [("0", "1"), ("1", "2"), ("2", "3")])
            _write_xy(same / "b.txt", [("0", "4"), ("1", "5"), ("2", "6")])
            _write_xy(diff / "a.txt", [("0", "1"), ("1", "2"), ("2", "3")])
            _write_xy(diff / "b.txt", [("10", "4"), ("11", "5"), ("12", "6")])
            same_xlsx, _csv = sto.export_spectra(
                list(same.glob("*.txt")),
                root / "same.xlsx",
                layout="auto",
            )
            diff_xlsx, _csv = sto.export_spectra(
                list(diff.glob("*.txt")),
                root / "diff.xlsx",
                layout="auto",
            )
            from openpyxl import load_workbook

            same_book = load_workbook(same_xlsx, data_only=True)
            same_sheet = same_book[same_book.sheetnames[0]]
            self.assertEqual(same_sheet.max_column, 3)
            diff_book = load_workbook(diff_xlsx, data_only=True)
            diff_sheet = diff_book[diff_book.sheetnames[0]]
            self.assertEqual(diff_sheet.max_column, 4)


def _origin_unavailable_path() -> Path:
    env = os.environ.get("ORIGIN_UNAVAILABLE_PATH", "").strip()
    if env:
        return Path(env)
    return Path(sto.__file__).resolve().parent / "origin_unavailable.txt"


def _record_origin_unavailable(reason: str) -> None:
    path = _origin_unavailable_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(reason, encoding="utf-8")


class BulkCliTests(unittest.TestCase):
    def test_cli_ingests_forty_plus_txts_into_one_workbook(self) -> None:
        n_files = 48
        tool = Path(sto.__file__).resolve()
        dest_dir = tool.parent / "_test_out"
        dest_dir.mkdir(parents=True, exist_ok=True)
        opju = dest_dir / "bulk.opju"
        if opju.exists():
            opju.unlink()
        with tempfile.TemporaryDirectory() as raw:
            folder = Path(raw) / "spectra"
            folder.mkdir()
            stems = []
            for index in range(n_files):
                stem = f"spec_{index:03d}"
                stems.append(stem)
                _write_xy(
                    folder / f"{stem}.txt",
                    [("0.0", str(index)), ("1.0", str(index + 1)), ("2.0", str(index + 2))],
                )
            self.assertNotEqual(Path(tempfile.gettempdir()).resolve().drive.lower(), opju.drive.lower())
            proc = subprocess.run(
                [
                    sys.executable,
                    str(tool),
                    "--cli",
                    "-i",
                    str(folder),
                    "-o",
                    str(opju),
                    "--layout",
                    "auto",
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=180,
            )
            if proc.returncode != 0 or not opju.exists() or opju.stat().st_size == 0:
                _record_origin_unavailable(
                    f"returncode={proc.returncode}\nstdout={proc.stdout}\nstderr={proc.stderr}\n"
                    f"opju_exists={opju.exists()}\n"
                )
                self.fail(f"Origin CLI did not write .opju; see origin_unavailable.txt\n{proc.stderr}")

            inspect = subprocess.run(
                [
                    sys.executable,
                    "-c",
                    (
                        "import sys\n"
                        f"sys.path.insert(0, r'{tool.parent.as_posix()}')\n"
                        "import originpro as op\n"
                        "import spectra_to_origin as sto\n"
                        "started = not sto.origin_process_running()\n"
                        "op.set_show(False)\n"
                        "assert op.open(sys.argv[1])\n"
                        "sheet = list(op.pages('w'))[0][0]\n"
                        "graphs = list(op.pages('g'))\n"
                        "n_cols = int(sheet.cols)\n"
                        "labels = [sheet.get_label(c, 'L') for c in range(n_cols)]\n"
                        "n_plots = len(list(graphs[0][0].obj.DataPlots)) if graphs else 0\n"
                        "print(n_cols, n_plots)\n"
                        "print(','.join(str(item) for item in labels))\n"
                        "sto.close_origin_app(op, started=started)\n"
                    ),
                    str(opju),
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=120,
            )
            if inspect.returncode != 0:
                _record_origin_unavailable(
                    f"inspect returncode={inspect.returncode}\nstdout={inspect.stdout}\nstderr={inspect.stderr}\n"
                )
                self.fail(f"Origin inspect failed; see origin_unavailable.txt\n{inspect.stderr}")
            lines = [line for line in inspect.stdout.splitlines() if line.strip()]
            cols_plots = lines[0].split()
            labels = lines[1].split(",")
            self.assertEqual(int(cols_plots[0]), 1 + n_files)
            self.assertEqual(int(cols_plots[1]), n_files)
            for stem in stems:
                self.assertIn(stem, labels)
            print(
                f"BULK_OPJU dest={opju} n_cols={cols_plots[0]} n_plots={cols_plots[1]}",
                flush=True,
            )

    def test_run_cli_origin_failure_is_visible(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            folder = Path(raw)
            _write_xy(folder / "a.txt", [("0", "1"), ("1", "2")])
            out = folder / "out.opju"
            args = sto.build_parser().parse_args(["--cli", "-i", str(folder), "-o", str(out)])

            def _boom(*_args, **_kwargs):
                raise sto.OriginExportError("boom")

            stderr = io.StringIO()
            with mock.patch.object(sto, "export_origin_project", side_effect=_boom):
                with contextlib.redirect_stderr(stderr):
                    code = sto.run_cli(args)
            self.assertEqual(code, 4)
            text = stderr.getvalue()
            self.assertIn("Origin 工程生成失败", text)
            self.assertIn("boom", text)


if __name__ == "__main__":
    unittest.main()
