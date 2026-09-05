"""Merge 2-column 1D spectra into an Origin .opju project (XYYY / XYXY)."""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import tkinter as tk
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

TEMP_COPY_DIR = "按温度"
HEADER_ROWS = 3
X_LONG_NAME = "2theta"
X_UNITS = "deg"
Y_UNITS = "a.u."

_HEADER_TEMP = re.compile(r"温度\s+([0-9.]+)\s*C")
_HEADER_TIME = re.compile(r"时效\s+([0-9.]+)\s*h")
_HEADER_ETA = re.compile(r"eta\s*=\s*([0-9.]+)", re.IGNORECASE)
_NAME_ETA = re.compile(r"_eta[0-9.]+$", re.IGNORECASE)
_NAME_TEMP = re.compile(r"(?<![0-9A-Za-z])(\d{2,4}C)(?![0-9A-Za-z])", re.IGNORECASE)
_TOKEN_SPLIT = re.compile(r"[_\-\s]+")
_SKIP_TOKEN = re.compile(
    r"^(?:eta[0-9.]+|\d{1,3}|t\d+\.?\d*h|\d+\.?\d*h)$",
    re.IGNORECASE,
)


class OriginExportError(RuntimeError):
    """Raised when Origin Pro did not produce a usable .opju file."""


@dataclass(frozen=True)
class Spectrum:
    path: Path
    x_text: tuple[str, ...]
    y_text: tuple[str, ...]
    long_name: str
    comment: str
    temperature_tag: str | None

    @property
    def n_points(self) -> int:
        return len(self.x_text)

    @property
    def x_values(self) -> list[float]:
        return [float(value) for value in self.x_text]

    @property
    def y_values(self) -> list[float]:
        return [float(value) for value in self.y_text]


@dataclass(frozen=True)
class SheetTable:
    name: str
    long_names: tuple[str, ...]
    units: tuple[str, ...]
    comments: tuple[str, ...]
    rows: tuple[tuple[str, ...], ...]


@dataclass(frozen=True)
class SheetSpec:
    name: str
    spectra: tuple[Spectrum, ...]
    layout: str


def read_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"无法解码：{path}")


def _long_name_from_stem(stem: str) -> str:
    return _NAME_ETA.sub("", stem)


def _temperature_tag(stem: str, header: str) -> str | None:
    match = _NAME_TEMP.search(stem)
    if match:
        return match.group(1).upper()
    match = _HEADER_TEMP.search(header)
    if match:
        return f"{match.group(1)}C"
    return None


def _comment_from_header(header: str, stem: str) -> str:
    parts: list[str] = []
    eta = _HEADER_ETA.search(header)
    temp = _HEADER_TEMP.search(header)
    time = _HEADER_TIME.search(header)
    if eta:
        parts.append(f"eta={eta.group(1)}")
    if temp:
        parts.append(f"{temp.group(1)} C")
    if time:
        parts.append(f"{time.group(1)} h")
    if not parts:
        parts.append(_long_name_from_stem(stem))
    return "; ".join(parts)


def parse_spectra(files: list[Path]) -> list[Spectrum]:
    return [parse_spectrum(path) for path in files]


def parse_spectrum(path: Path) -> Spectrum:
    text = read_text(path)
    header_lines: list[str] = []
    x_text: list[str] = []
    y_text: list[str] = []
    for line_number, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("#"):
            header_lines.append(line.lstrip("# ").strip())
            continue
        parts = line.split()
        if len(parts) != 2:
            raise ValueError(f"{path.name} 第 {line_number} 行不是两列：{raw_line}")
        try:
            float(parts[0])
            float(parts[1])
        except ValueError as exc:
            raise ValueError(f"{path.name} 第 {line_number} 行不是数值：{raw_line}") from exc
        x_text.append(parts[0])
        y_text.append(parts[1])
    if len(x_text) < 2:
        raise ValueError(f"{path.name} 有效数据点不足：{len(x_text)}")
    header = "    ".join(header_lines)
    stem = path.stem
    return Spectrum(
        path=path,
        x_text=tuple(x_text),
        y_text=tuple(y_text),
        long_name=_long_name_from_stem(stem),
        comment=_comment_from_header(header, stem),
        temperature_tag=_temperature_tag(stem, header),
    )


def under_temp_copy_dir(path: Path) -> bool:
    return TEMP_COPY_DIR in path.parts


def collect_txt_files(root: Path) -> list[Path]:
    if root.is_file():
        if root.suffix.lower() != ".txt":
            raise ValueError(f"不是 txt：{root}")
        return [root.resolve()]
    if not root.is_dir():
        raise ValueError(f"路径不存在：{root}")

    found: list[Path] = []
    for dirpath, dirnames, filenames in root.walk():
        txt_here = [dirpath / name for name in filenames if name.lower().endswith(".txt")]
        if txt_here and TEMP_COPY_DIR in dirnames:
            dirnames.remove(TEMP_COPY_DIR)
        found.extend(path.resolve() for path in txt_here)
    return found


def merge_file_list(existing: list[Path], incoming: list[Path]) -> list[Path]:
    by_name = {path.name.lower(): path for path in existing}
    out = list(existing)
    seen = {path.resolve() for path in existing}
    for path in incoming:
        resolved = path.resolve()
        if resolved in seen:
            continue
        key = path.name.lower()
        old = by_name.get(key)
        if old is not None:
            if under_temp_copy_dir(old) and not under_temp_copy_dir(path):
                out = [path if item == old else item for item in out]
                seen.discard(old.resolve())
                seen.add(resolved)
                by_name[key] = path
            continue
        out.append(path)
        seen.add(resolved)
        by_name[key] = path
    return out


def collect_from_user_paths(paths: list[Path]) -> tuple[list[Path], list[str]]:
    incoming: list[Path] = []
    errors: list[str] = []
    for path in paths:
        try:
            incoming.extend(collect_txt_files(path))
        except ValueError as exc:
            errors.append(str(exc))
    return incoming, errors


def parse_drop_paths(payload: str | bytes | list[str] | list[Path]) -> list[Path]:
    if isinstance(payload, (list, tuple)):
        return [Path(str(item).strip().strip('"')) for item in payload if str(item).strip()]
    text = payload.decode("utf-16le") if isinstance(payload, (bytes, bytearray)) else str(payload)
    parts = re.split(r"[\r\n\x00]+", text.strip().strip("\x00"))
    return [Path(part.strip().strip('"')) for part in parts if part.strip()]


def load_from_drop_payload(existing: list[Path], payload: str | bytes | list[str] | list[Path]) -> list[Path]:
    incoming, _errors = collect_from_user_paths(parse_drop_paths(payload))
    return merge_file_list(existing, incoming)


def shared_x_grid(spectra: list[Spectrum]) -> bool:
    if not spectra:
        return False
    first = spectra[0].x_text
    return all(item.x_text == first for item in spectra[1:])


def infer_layout(spectra: list[Spectrum]) -> str:
    if shared_x_grid(spectra):
        return "XYYY"
    return "XYXY"


def resolve_layout(spectra: list[Spectrum], layout: str) -> str:
    text = (layout or "auto").strip().upper()
    if text in {"", "AUTO"}:
        return infer_layout(spectra)
    if text in {"XYYY", "XYXY"}:
        return text
    raise ValueError(f"不支持的布局：{layout}")


def _xyyy_table(name: str, spectra: list[Spectrum]) -> SheetTable:
    if not shared_x_grid(spectra):
        names = ", ".join(item.path.name for item in spectra[:4])
        raise ValueError(f"X 网格不一致，不能用 XYYY。请改选 XYXY。涉及文件：{names}")
    long_names = (X_LONG_NAME, *(item.long_name for item in spectra))
    units = (X_UNITS, *(Y_UNITS for _ in spectra))
    comments = ("", *(item.comment for item in spectra))
    rows = []
    x_text = spectra[0].x_text
    for index, x_value in enumerate(x_text):
        rows.append((x_value, *(item.y_text[index] for item in spectra)))
    return SheetTable(name, long_names, units, comments, tuple(rows))


def _xyxy_table(name: str, spectra: list[Spectrum]) -> SheetTable:
    long_names: list[str] = []
    units: list[str] = []
    comments: list[str] = []
    for item in spectra:
        long_names.extend([X_LONG_NAME, item.long_name])
        units.extend([X_UNITS, Y_UNITS])
        comments.extend(["", item.comment])
    n_rows = max(item.n_points for item in spectra)
    rows: list[tuple[str, ...]] = []
    for index in range(n_rows):
        cells: list[str] = []
        for item in spectra:
            if index < item.n_points:
                cells.extend([item.x_text[index], item.y_text[index]])
            else:
                cells.extend(["", ""])
        rows.append(tuple(cells))
    return SheetTable(name, tuple(long_names), tuple(units), tuple(comments), tuple(rows))


def build_table(name: str, spectra: list[Spectrum], layout: str) -> SheetTable:
    if not spectra:
        raise ValueError("没有谱线")
    layout = resolve_layout(spectra, layout)
    if layout == "XYYY":
        return _xyyy_table(name, spectra)
    if layout == "XYXY":
        return _xyxy_table(name, spectra)
    raise ValueError(f"不支持的布局：{layout}")


def _stem_tokens(stem: str) -> list[str]:
    parts = [part for part in _TOKEN_SPLIT.split(stem) if part]
    for match in _NAME_TEMP.finditer(stem):
        token = match.group(1)
        if token not in parts:
            parts.append(token)
    return parts


def _usable_token(token: str) -> bool:
    return not _SKIP_TOKEN.match(token)


def _score_index_groups(groups: list[tuple[str, list[int]]], sample: str) -> int:
    sizes = [len(indexes) for _name, indexes in groups]
    n_items = sum(sizes)
    n_groups = len(groups)
    if n_items == 0 or n_groups < 2:
        return -10_000
    score = 0
    if any(name == "other" for name, _indexes in groups):
        other_n = next(len(indexes) for name, indexes in groups if name == "other")
        score -= 50 * other_n
    else:
        score += 25
    if min(sizes) >= 2:
        score += 20
    score -= 8 * sum(1 for size in sizes if size == 1)
    if n_groups <= max(2, n_items // 2):
        score += 12
    else:
        score -= 12
    if any(char.isalpha() for char in sample):
        score += 10
    score += min(n_groups, 8)
    mean = n_items / n_groups
    score -= int(sum(abs(size - mean) for size in sizes))
    return score


def _groups_from_assignment(assignment: list[str | None]) -> list[tuple[str, list[int]]]:
    order: list[str] = []
    buckets: dict[str, list[int]] = {}
    other: list[int] = []
    for index, name in enumerate(assignment):
        if name is None:
            other.append(index)
            continue
        if name not in buckets:
            order.append(name)
            buckets[name] = []
        buckets[name].append(index)
    groups = [(name, buckets[name]) for name in order]
    if other:
        groups.append(("other", other))
    return groups


def suggest_filename_groups(stems: list[str]) -> list[tuple[str, list[int]]]:
    n_items = len(stems)
    if n_items == 0:
        return []
    if n_items == 1:
        return [("all", [0])]

    token_lists = [_stem_tokens(stem) for stem in stems]
    candidates: list[tuple[int, list[tuple[str, list[int]]]]] = []

    lengths = {len(tokens) for tokens in token_lists}
    if len(lengths) == 1:
        width = next(iter(lengths))
        for column in range(width):
            values = [token_lists[index][column] for index in range(n_items)]
            if all(not _usable_token(value) for value in values):
                continue
            unique = list(dict.fromkeys(values))
            if len(unique) < 2 or len(unique) == n_items:
                continue
            groups = _groups_from_assignment(list(values))
            candidates.append((_score_index_groups(groups, values[0]), groups))

    freq: Counter[str] = Counter()
    for tokens in token_lists:
        for token in dict.fromkeys(tokens):
            if _usable_token(token):
                freq[token] += 1
    covering = {token for token, count in freq.items() if 2 <= count < n_items}
    if covering:
        assignment: list[str | None] = []
        for tokens in token_lists:
            found = [token for token in tokens if token in covering]
            assignment.append(found[0] if found else None)
        groups = _groups_from_assignment(assignment)
        named = [name for name, _indexes in groups if name != "other"]
        if len(named) >= 2:
            sample = named[0]
            candidates.append((_score_index_groups(groups, sample), groups))

    if not candidates:
        return [("all", list(range(n_items)))]
    candidates.sort(key=lambda item: item[0], reverse=True)
    best = candidates[0][1]
    named = [name for name, _indexes in best if name != "other"]
    if len(named) < 2:
        return [("all", list(range(n_items)))]
    return best


def groups_from_filenames(spectra: list[Spectrum]) -> list[tuple[str, list[Spectrum]]]:
    grouped = suggest_filename_groups([item.path.stem for item in spectra])
    return [(name, [spectra[index] for index in indexes]) for name, indexes in grouped]


def partition_even(n_items: int, n_groups: int, prefix: str = "group") -> list[tuple[str, list[int]]]:
    if n_items <= 0:
        return []
    if n_groups < 1:
        raise ValueError("分组数至少为 1")
    n_groups = min(n_groups, n_items)
    base, extra = divmod(n_items, n_groups)
    groups: list[tuple[str, list[int]]] = []
    start = 0
    for index in range(n_groups):
        size = base + (1 if index < extra else 0)
        end = start + size
        groups.append((f"{prefix}{index + 1}", list(range(start, end))))
        start = end
    return groups


def partition_spectra(spectra: list[Spectrum], n_groups: int, prefix: str = "group") -> list[tuple[str, list[Spectrum]]]:
    return [
        (name, [spectra[index] for index in indexes])
        for name, indexes in partition_even(len(spectra), n_groups, prefix)
    ]


def assignments_from_groups(n_items: int, groups: list[tuple[str, list[int]]]) -> list[str]:
    assigned = ["all"] * n_items
    for name, indexes in groups:
        for index in indexes:
            assigned[index] = name
    return assigned


def groups_from_assignments(items: list, assignments: list[str]) -> list[tuple[str, list]]:
    if len(items) != len(assignments):
        raise ValueError("分组标记与谱线数量不一致")
    order: list[str] = []
    buckets: dict[str, list] = {}
    for item, name in zip(items, assignments):
        label = name.strip() or "group"
        if label not in buckets:
            order.append(label)
            buckets[label] = []
        buckets[label].append(item)
    return [(name, buckets[name]) for name in order]


def resolve_groups(
    spectra: list[Spectrum],
    groups: list[tuple[str, list[Spectrum]]] | None = None,
    split_temp: bool = False,
    n_groups: int | None = None,
) -> list[tuple[str, list[Spectrum]]]:
    if groups is not None:
        return [(name, list(items)) for name, items in groups if items]
    if n_groups is not None and n_groups > 1:
        return partition_spectra(spectra, n_groups)
    if split_temp:
        return groups_from_filenames(spectra)
    return [("spectra", list(spectra))]


def build_sheet_specs(
    groups: list[tuple[str, list[Spectrum]]],
    layout: str,
) -> list[SheetSpec]:
    if not groups:
        raise ValueError("没有分组")
    all_spectra = [item for _name, items in groups for item in items]
    if not all_spectra:
        raise ValueError("没有谱线")
    resolved = resolve_layout(all_spectra, layout)
    specs: list[SheetSpec] = []
    for name, items in groups:
        if not items:
            continue
        specs.append(SheetSpec(name=name, spectra=tuple(items), layout=resolved))
    if not specs:
        raise ValueError("没有可导出的 sheet")
    return specs


def group_by_temperature(spectra: list[Spectrum]) -> dict[str, list[Spectrum]]:
    grouped: dict[str, list[Spectrum]] = {}
    for item in spectra:
        tag = item.temperature_tag or "other"
        grouped.setdefault(tag, []).append(item)
    return grouped


def sheet_name(raw: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", raw).strip() or "sheet"
    return cleaned[:31]


def origin_sheet_name(raw: str) -> str:
    token = re.sub(r"[^0-9A-Za-z]+", "", str(raw).strip()) or "sheet"
    if token[0].isdigit():
        token = "T" + token
    return token[:13]


def unique_origin_names(raw_names: list[str], limit: int = 13) -> list[str]:
    used: set[str] = set()
    names: list[str] = []
    for raw in raw_names:
        base = origin_sheet_name(raw)[:limit]
        candidate = base
        serial = 2
        while candidate.lower() in used:
            suffix = str(serial)
            candidate = (base[: max(1, limit - len(suffix))] + suffix)[:limit]
            serial += 1
        used.add(candidate.lower())
        names.append(candidate)
    return names


def readme_rows(spectra: list[Spectrum], layout: str, group_names: list[str]) -> list[list[str]]:
    shared = "yes" if shared_x_grid(spectra) else "no"
    rows = [
        ["layout", layout],
        ["n_spectra", str(len(spectra))],
        ["n_points_first", str(spectra[0].n_points if spectra else 0)],
        ["shared_x", shared],
        ["groups", ", ".join(group_names)],
        ["header_rows", "Long Name / Units / Comments"],
        [],
        ["index", "long_name", "n_points", "group_hint", "comment", "path"],
    ]
    for index, item in enumerate(spectra, start=1):
        rows.append(
            [
                str(index),
                item.long_name,
                str(item.n_points),
                item.temperature_tag or "",
                item.comment,
                str(item.path),
            ]
        )
    return rows


def _write_data_sheet(workbook: Workbook, table: SheetTable) -> None:
    sheet = workbook.create_sheet(sheet_name(table.name))
    sheet.append(list(table.long_names))
    sheet.append(list(table.units))
    sheet.append(list(table.comments))
    for row in table.rows:
        sheet.append([_excel_value(cell) for cell in row])
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A4"
    for index, name in enumerate(table.long_names, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(36, max(12, len(name) + 2))


def _excel_value(cell: str):
    if cell == "":
        return None
    return float(cell)


def write_xlsx(
    path: Path,
    tables: list[SheetTable],
    spectra: list[Spectrum],
    layout: str,
    group_names: list[str],
) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for table in tables:
        _write_data_sheet(workbook, table)
    readme = workbook.create_sheet("readme")
    for row in readme_rows(spectra, layout, group_names):
        readme.append(row)
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 28
    readme.column_dimensions["F"].width = 80
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_csv(path: Path, table: SheetTable) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(table.long_names)
        writer.writerow(table.units)
        writer.writerow(table.comments)
        writer.writerows(table.rows)


def build_export_tables(spectra: list[Spectrum], layout: str, groups: list[tuple[str, list[Spectrum]]]) -> list[SheetTable]:
    specs = build_sheet_specs(groups, layout)
    return [build_table(spec.name, list(spec.spectra), spec.layout) for spec in specs]


def export_spectra(
    files: list[Path],
    output_xlsx: Path,
    layout: str = "auto",
    split_temp: bool = False,
    n_groups: int | None = None,
    groups: list[tuple[str, list[Spectrum]]] | None = None,
    write_csv_copy: bool = True,
) -> tuple[Path, Path | None]:
    if not files:
        raise ValueError("没有谱线文件")
    if groups:
        spectra = [item for _name, items in groups for item in items]
    else:
        spectra = parse_spectra(files)
    if not spectra:
        raise ValueError("没有谱线文件")
    resolved_groups = resolve_groups(spectra, groups=groups, split_temp=split_temp, n_groups=n_groups)
    resolved_layout = resolve_layout(spectra, layout)
    tables = build_export_tables(spectra, resolved_layout, resolved_groups)
    write_xlsx(
        output_xlsx,
        tables,
        spectra,
        resolved_layout,
        [name for name, _items in resolved_groups],
    )
    csv_path = None
    if write_csv_copy:
        csv_path = output_xlsx.with_suffix(".csv")
        write_csv(csv_path, tables[0])
    return output_xlsx, csv_path


def origin_process_running() -> bool:
    if sys.platform != "win32":
        return False
    try:
        output = subprocess.check_output(
            ["tasklist", "/FI", "IMAGENAME eq Origin64.exe", "/NH"],
            text=True,
            errors="ignore",
        )
    except (OSError, subprocess.CalledProcessError):
        return False
    return "Origin64.exe" in output


def close_origin_app(op, *, started: bool, timeout: float = 8.0) -> None:
    """Quit Origin COM and, if this call started the process, wait until it dies."""
    try:
        op.exit()
    except Exception:
        pass
    if not started or sys.platform != "win32":
        return
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        if not origin_process_running():
            return
        time.sleep(0.2)
    if origin_process_running():
        subprocess.run(
            ["taskkill", "/F", "/IM", "Origin64.exe", "/T"],
            capture_output=True,
            text=True,
            errors="ignore",
        )


def _load_originpro():
    try:
        import originpro as op
    except ImportError as exc:
        raise OriginExportError(
            "当前 Python 无法 import originpro，不能生成 .opju。"
            "请在已安装 Origin Pro 的环境里运行，并确认 originpro / OriginExt 可用。"
        ) from exc
    return op


def _pad_column(values: list[float], n_rows: int) -> list[float]:
    if len(values) >= n_rows:
        return values[:n_rows]
    return values + [float("nan")] * (n_rows - len(values))


def _fill_origin_sheet(sheet, spectra: list[Spectrum], layout: str) -> None:
    layout = resolve_layout(spectra, layout)
    if layout == "XYYY":
        if not shared_x_grid(spectra):
            names = ", ".join(item.path.name for item in spectra[:4])
            raise ValueError(f"X 网格不一致，不能用 XYYY。请改选 XYXY。涉及文件：{names}")
        columns = [spectra[0].x_values] + [item.y_values for item in spectra]
        longs = [X_LONG_NAME, *(item.long_name for item in spectra)]
        units = [X_UNITS, *(Y_UNITS for _ in spectra)]
        comments = ["", *(item.comment for item in spectra)]
        axes = "x" + "y" * len(spectra)
    elif layout == "XYXY":
        n_rows = max(item.n_points for item in spectra)
        columns = []
        longs: list[str] = []
        units: list[str] = []
        comments: list[str] = []
        for item in spectra:
            columns.append(_pad_column(item.x_values, n_rows))
            columns.append(_pad_column(item.y_values, n_rows))
            longs.extend([X_LONG_NAME, item.long_name])
            units.extend([X_UNITS, Y_UNITS])
            comments.extend(["", item.comment])
        axes = "xy" * len(spectra)
    else:
        raise ValueError(f"不支持的布局：{layout}")
    sheet.cols = len(columns)
    sheet.from_list2(columns, 0, 0)
    sheet.set_labels(longs, "L")
    sheet.set_labels(units, "U")
    sheet.set_labels(comments, "C")
    sheet.cols_axis(axes)


def _plot_origin_sheet(op, sheet, spectra: list[Spectrum], layout: str, graph_name: str):
    graph = op.new_graph(lname=graph_name, hidden=False)
    if graph is None:
        raise OriginExportError(f"Origin 未能创建图：{graph_name}")
    layer = graph[0]
    layout = resolve_layout(spectra, layout)
    if layout == "XYYY":
        for y_col in range(1, len(spectra) + 1):
            plot = layer.add_plot(sheet, coly=y_col, colx=0, type="l")
            if plot is None:
                raise OriginExportError(f"Origin 未能把第 {y_col} 列画进 {graph_name}")
    else:
        for index in range(len(spectra)):
            x_col = 2 * index
            y_col = x_col + 1
            plot = layer.add_plot(sheet, coly=y_col, colx=x_col, type="l")
            if plot is None:
                raise OriginExportError(f"Origin 未能把 {spectra[index].long_name} 画进 {graph_name}")
    n_plots = len(list(layer.obj.DataPlots))
    if n_plots != len(spectra):
        raise OriginExportError(
            f"{graph_name} 应有 {len(spectra)} 条曲线，Origin 实际只画了 {n_plots} 条。"
        )
    layer.axis("x").title = f"{X_LONG_NAME} ({X_UNITS})"
    layer.axis("y").title = f"Intensity ({Y_UNITS})"
    layer.rescale()
    layer.lt_exec("legend -r")
    return graph


def _save_origin_project(op, path: Path) -> Path:
    path = path.expanduser().resolve()
    if path.suffix.lower() != ".opju":
        path = path.with_suffix(".opju")
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_dir = Path(tempfile.mkdtemp(prefix="spectra_opju_"))
    tmp_path = tmp_dir / "export.opju"
    saved = op.save(str(tmp_path))
    if not tmp_path.exists() or tmp_path.stat().st_size == 0:
        saved = op.save(str(path))
        shutil.rmtree(tmp_dir, ignore_errors=True)
        if not path.exists() or path.stat().st_size == 0:
            raise OriginExportError(
                f"Origin 没有写出 .opju 文件（save 返回 {saved!r}）。"
                "请确认 Origin Pro 已启动且未被其他对话框挡住。"
            )
        return path
    os.replace(tmp_path, path)
    shutil.rmtree(tmp_dir, ignore_errors=True)
    if not path.exists() or path.stat().st_size == 0:
        raise OriginExportError(f"Origin 保存后文件仍不存在或为空：{path}")
    return path


def write_origin_project(
    spectra: list[Spectrum],
    output_opju: Path,
    layout: str = "auto",
    split_temp: bool = False,
    n_groups: int | None = None,
    groups: list[tuple[str, list[Spectrum]]] | None = None,
    show_origin: bool = False,
    keep_open: bool = False,
) -> Path:
    if not spectra:
        raise ValueError("没有谱线")
    resolved_groups = resolve_groups(spectra, groups=groups, split_temp=split_temp, n_groups=n_groups)
    specs = build_sheet_specs(resolved_groups, layout)
    resolved_layout = specs[0].layout
    op = _load_originpro()
    closed = False
    started = not origin_process_running()
    try:
        op.set_show(bool(show_origin) or bool(keep_open))
        op.new(False)
        books = list(op.pages("w"))
        if not books:
            raise OriginExportError("Origin 新建工程后没有工作簿。")
        book = books[0]
        book.lname = f"spectra_{resolved_layout}"
        for extra in books[1:]:
            extra.destroy()
        shorts = unique_origin_names([spec.name for spec in specs])
        for index, (spec, short) in enumerate(zip(specs, shorts)):
            sheet = book[0] if index == 0 else book.add_sheet(short)
            sheet.name = short
            sheet.lname = spec.name
            items = list(spec.spectra)
            _fill_origin_sheet(sheet, items, spec.layout)
            _plot_origin_sheet(op, sheet, items, spec.layout, f"{short}_plot")
        saved = _save_origin_project(op, output_opju)
        if keep_open:
            op.set_show(True)
        else:
            close_origin_app(op, started=started)
            closed = True
        if not saved.exists() or saved.stat().st_size == 0:
            raise OriginExportError(f"Origin 保存后文件仍不存在或为空：{saved}")
        return saved
    except OriginExportError:
        raise
    except Exception as exc:
        raise OriginExportError(f"Origin 工程生成失败：{exc}") from exc
    finally:
        if not keep_open and not closed:
            close_origin_app(op, started=started)


def export_origin_project(
    files: list[Path],
    output_opju: Path,
    layout: str = "auto",
    split_temp: bool = False,
    n_groups: int | None = None,
    groups: list[tuple[str, list[Spectrum]]] | None = None,
    show_origin: bool = False,
    keep_open: bool = False,
    also_xlsx: bool = False,
) -> tuple[Path, Path | None, Path | None]:
    if not files and not groups:
        raise ValueError("没有谱线文件")
    if groups:
        spectra = [item for _name, items in groups for item in items]
    else:
        spectra = parse_spectra(files)
    if not spectra:
        raise ValueError("没有谱线文件")
    resolved_groups = resolve_groups(spectra, groups=groups, split_temp=split_temp, n_groups=n_groups)
    opju_path = write_origin_project(
        spectra,
        output_opju,
        layout=layout,
        groups=resolved_groups,
        show_origin=show_origin,
        keep_open=keep_open,
    )
    xlsx_path = None
    csv_path = None
    if also_xlsx:
        xlsx_path, csv_path = export_spectra(
            files,
            opju_path.with_suffix(".xlsx"),
            layout=layout,
            groups=resolved_groups,
            write_csv_copy=True,
        )
    return opju_path, xlsx_path, csv_path


def enable_windows_file_drop(widget: tk.Misc, on_paths) -> bool:
    """Hook WM_DROPFILES so Explorer can drop files or folders onto the window."""
    if sys.platform != "win32":
        return False
    try:
        return _enable_windows_file_drop(widget, on_paths)
    except Exception:
        return False


def _enable_windows_file_drop(widget: tk.Misc, on_paths) -> bool:
    import ctypes
    from ctypes import wintypes

    WM_DROPFILES = 0x0233
    GWLP_WNDPROC = -4
    GA_ROOT = 2
    LRESULT = ctypes.c_int64 if ctypes.sizeof(ctypes.c_void_p) == 8 else ctypes.c_long
    user32 = ctypes.windll.user32
    shell32 = ctypes.windll.shell32

    user32.GetParent.argtypes = [wintypes.HWND]
    user32.GetParent.restype = wintypes.HWND
    user32.GetAncestor.argtypes = [wintypes.HWND, ctypes.c_uint]
    user32.GetAncestor.restype = wintypes.HWND
    DragAcceptFiles = shell32.DragAcceptFiles
    DragAcceptFiles.argtypes = [wintypes.HWND, wintypes.BOOL]
    DragQueryFileW = shell32.DragQueryFileW
    DragQueryFileW.argtypes = [wintypes.HANDLE, wintypes.UINT, wintypes.LPWSTR, wintypes.UINT]
    DragQueryFileW.restype = wintypes.UINT
    DragFinish = shell32.DragFinish
    DragFinish.argtypes = [wintypes.HANDLE]

    widget.update_idletasks()
    hwnd = wintypes.HWND(int(widget.winfo_id()))
    parent = user32.GetParent(hwnd)
    if parent:
        hwnd = parent
    root = user32.GetAncestor(hwnd, GA_ROOT)
    if root:
        hwnd = root

    WNDPROC = ctypes.WINFUNCTYPE(LRESULT, wintypes.HWND, wintypes.UINT, wintypes.WPARAM, wintypes.LPARAM)
    GetWindowLongPtr = user32.GetWindowLongPtrW
    SetWindowLongPtr = user32.SetWindowLongPtrW
    CallWindowProc = user32.CallWindowProcW
    GetWindowLongPtr.argtypes = [wintypes.HWND, ctypes.c_int]
    GetWindowLongPtr.restype = ctypes.c_void_p
    SetWindowLongPtr.argtypes = [wintypes.HWND, ctypes.c_int, ctypes.c_void_p]
    SetWindowLongPtr.restype = ctypes.c_void_p
    CallWindowProc.argtypes = [ctypes.c_void_p, wintypes.HWND, wintypes.UINT, wintypes.WPARAM, wintypes.LPARAM]
    CallWindowProc.restype = LRESULT

    old_proc = GetWindowLongPtr(hwnd, GWLP_WNDPROC)
    if not old_proc:
        return False

    def _paths_from_hdrop(hdrop) -> list[str]:
        count = DragQueryFileW(hdrop, 0xFFFFFFFF, None, 0)
        paths: list[str] = []
        for index in range(count):
            length = DragQueryFileW(hdrop, index, None, 0) + 1
            buffer = ctypes.create_unicode_buffer(length)
            DragQueryFileW(hdrop, index, buffer, length)
            paths.append(buffer.value)
        DragFinish(hdrop)
        return paths

    def _wndproc(hwnd_value, message, wparam, lparam):
        if message == WM_DROPFILES:
            try:
                paths = _paths_from_hdrop(wparam)
                widget.after(0, lambda: on_paths(paths))
            except Exception:
                pass
            return 0
        return CallWindowProc(old_proc, hwnd_value, message, wparam, lparam)

    new_proc = WNDPROC(_wndproc)
    widget._drop_wndproc = new_proc  # noqa: SLF001 — keep callback alive
    widget._drop_oldproc = old_proc  # noqa: SLF001
    widget._drop_hwnd = hwnd  # noqa: SLF001

    def _restore(_event=None):
        try:
            SetWindowLongPtr(hwnd, GWLP_WNDPROC, old_proc)
        except Exception:
            pass

    widget.bind("<Destroy>", _restore, add="+")
    SetWindowLongPtr(hwnd, GWLP_WNDPROC, ctypes.cast(new_proc, ctypes.c_void_p))
    DragAcceptFiles(hwnd, True)
    return True


def _status_text(spectra: list[Spectrum], layout: str, assignments: list[str]) -> str:
    if not spectra:
        return "把谱线 txt 或文件夹拖进窗口，或点添加。X 相同会识别为 XYYY，不同则为 XYXY。"
    inferred = infer_layout(spectra)
    resolved = resolve_layout(spectra, layout)
    n_points = {item.n_points for item in spectra}
    points_text = str(next(iter(n_points))) if len(n_points) == 1 else "点数不同"
    shared = "X 相同" if shared_x_grid(spectra) else "X 不同"
    n_cols = 1 + len(spectra) if resolved == "XYYY" else 2 * len(spectra)
    group_names = list(dict.fromkeys(assignments)) if assignments else ["all"]
    auto = f"；自动识别 {inferred}" if (layout or "auto").strip().upper() in {"", "AUTO"} else ""
    warn = ""
    if resolved == "XYYY" and not shared_x_grid(spectra):
        warn = " | XYYY 需要相同 X，请改 XYXY 或自动"
    return (
        f"{len(spectra)} 条 | {points_text} | {shared} | {resolved} → {n_cols} 列"
        f"{auto} | {len(group_names)} 张 sheet：{', '.join(group_names)}{warn}"
    )


class SpectraToOriginApp:
    def __init__(self, initial_files: list[Path] | None = None) -> None:
        self.files: list[Path] = []
        self.assignments: list[str] = []
        self._user_grouped = False
        self._cached_key: tuple[Path, ...] | None = None
        self._cached_spectra: list[Spectrum] = []
        self.root = tk.Tk()
        self.root.title("谱线 → Origin 工程")
        self.root.minsize(820, 560)
        self.layout_var = tk.StringVar(value="auto")
        self.n_groups_var = tk.StringVar(value="3")
        self.keep_open_var = tk.BooleanVar(value=False)
        self.show_origin_var = tk.BooleanVar(value=False)
        self.also_xlsx_var = tk.BooleanVar(value=False)
        self.group_choice = tk.StringVar(value="")
        self.rename_var = tk.StringVar(value="")
        self._build()
        self.root.after(200, self._hook_drop)
        if initial_files:
            self._add_paths(initial_files)

    def _build(self) -> None:
        frame = ttk.Frame(self.root, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        hint = ttk.Label(frame, text="把谱线文件或文件夹拖到这个窗口即可加载")
        hint.pack(fill=tk.X)

        list_frame = ttk.Frame(frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(6, 0))
        scroll = ttk.Scrollbar(list_frame)
        self.listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED, height=16)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.config(yscrollcommand=scroll.set)
        scroll.config(command=self.listbox.yview)

        buttons = ttk.Frame(frame)
        buttons.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(buttons, text="添加文件", command=self.add_files).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(buttons, text="添加文件夹", command=self.add_folder).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(buttons, text="删除选中", command=self.remove_selected).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(buttons, text="上移", command=lambda: self.move_selected(-1)).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(buttons, text="下移", command=lambda: self.move_selected(1)).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(buttons, text="清空", command=self.clear_files).pack(side=tk.LEFT)

        options = ttk.Frame(frame)
        options.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(options, text="布局").pack(side=tk.LEFT)
        ttk.Radiobutton(
            options, text="自动", value="auto", variable=self.layout_var, command=self._refresh_status
        ).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Radiobutton(
            options, text="XYYY（共用 X）", value="XYYY", variable=self.layout_var, command=self._refresh_status
        ).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Radiobutton(
            options, text="XYXY（每条自己的 X）", value="XYXY", variable=self.layout_var, command=self._refresh_status
        ).pack(side=tk.LEFT, padx=(8, 0))

        groups = ttk.Frame(frame)
        groups.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(groups, text="按文件名分组", command=self.apply_filename_groups).pack(side=tk.LEFT)
        ttk.Label(groups, text="均分成").pack(side=tk.LEFT, padx=(12, 4))
        ttk.Spinbox(groups, from_=2, to=12, width=4, textvariable=self.n_groups_var).pack(side=tk.LEFT)
        ttk.Label(groups, text="张").pack(side=tk.LEFT, padx=(4, 6))
        ttk.Button(groups, text="应用均分", command=self.apply_even_split).pack(side=tk.LEFT)
        ttk.Button(groups, text="全部一张表", command=self.apply_single_sheet).pack(side=tk.LEFT, padx=(8, 0))

        move = ttk.Frame(frame)
        move.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(move, text="选中移入").pack(side=tk.LEFT)
        self.group_combo = ttk.Combobox(move, textvariable=self.group_choice, width=14, state="readonly")
        self.group_combo.pack(side=tk.LEFT, padx=(6, 6))
        ttk.Button(move, text="移入", command=self.move_selected_to_group).pack(side=tk.LEFT)
        ttk.Label(move, text="组改名").pack(side=tk.LEFT, padx=(12, 4))
        ttk.Entry(move, textvariable=self.rename_var, width=16).pack(side=tk.LEFT)
        ttk.Button(move, text="改名", command=self.rename_selected_group).pack(side=tk.LEFT, padx=(6, 0))

        origin_opts = ttk.Frame(frame)
        origin_opts.pack(fill=tk.X, pady=(8, 0))
        ttk.Checkbutton(origin_opts, text="完成后保持 Origin 打开", variable=self.keep_open_var).pack(side=tk.LEFT)
        ttk.Checkbutton(origin_opts, text="显示 Origin 窗口", variable=self.show_origin_var).pack(side=tk.LEFT, padx=(16, 0))
        ttk.Checkbutton(origin_opts, text="同时导出 xlsx/csv", variable=self.also_xlsx_var).pack(side=tk.LEFT, padx=(16, 0))

        export_row = ttk.Frame(frame)
        export_row.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(export_row, text="生成 Origin 工程 (.opju)", command=self.export).pack(side=tk.LEFT)

        self.status = ttk.Label(frame, text="", wraplength=780, justify=tk.LEFT)
        self.status.pack(fill=tk.X, pady=(10, 0))
        self._refresh_status()

    def _hook_drop(self) -> None:
        enable_windows_file_drop(self.root, self._on_drop_paths)

    def _on_drop_paths(self, paths: list[str]) -> None:
        self._add_paths([Path(item) for item in paths])

    def _refresh_list(self) -> None:
        self.listbox.delete(0, tk.END)
        for path, group in zip(self.files, self.assignments):
            self.listbox.insert(tk.END, f"{group}  |  {path.name}")
        names = list(dict.fromkeys(self.assignments))
        self.group_combo["values"] = names
        if names and self.group_choice.get() not in names:
            self.group_choice.set(names[0])
        self._refresh_status()

    def _parsed_spectra(self) -> list[Spectrum]:
        key = tuple(self.files)
        if self._cached_key != key:
            self._cached_spectra = parse_spectra(self.files) if self.files else []
            self._cached_key = key
        return self._cached_spectra

    def _refresh_status(self) -> None:
        try:
            spectra = self._parsed_spectra()
        except ValueError as exc:
            self.status.config(text=f"读取失败：{exc}")
            return
        self.status.config(text=_status_text(spectra, self.layout_var.get(), self.assignments))

    def _set_assignments(self, assignments: list[str], *, user: bool) -> None:
        self.assignments = assignments
        if user:
            self._user_grouped = True
        self._refresh_list()

    def _auto_assign(self) -> None:
        if not self.files:
            self.assignments = []
            return
        grouped = suggest_filename_groups([path.stem for path in self.files])
        self.assignments = assignments_from_groups(len(self.files), grouped)

    def _add_paths(self, paths: list[Path]) -> None:
        _incoming, errors = collect_from_user_paths(paths)
        self.files = load_from_drop_payload(self.files, paths)
        if not self._user_grouped:
            self._auto_assign()
        else:
            while len(self.assignments) < len(self.files):
                self.assignments.append(self.assignments[-1] if self.assignments else "all")
            self.assignments = self.assignments[: len(self.files)]
        self._refresh_list()
        if errors:
            messagebox.showwarning("部分路径未加入", "\n".join(errors[:8]))

    def add_files(self) -> None:
        selected = filedialog.askopenfilenames(
            title="选择谱线 txt",
            filetypes=[("谱线文本", "*.txt"), ("所有文件", "*.*")],
        )
        if selected:
            self._add_paths([Path(item) for item in selected])

    def add_folder(self) -> None:
        selected = filedialog.askdirectory(title="选择谱线文件夹")
        if selected:
            self._add_paths([Path(selected)])

    def remove_selected(self) -> None:
        indexes = sorted(self.listbox.curselection(), reverse=True)
        for index in indexes:
            del self.files[index]
            del self.assignments[index]
        self._refresh_list()

    def move_selected(self, step: int) -> None:
        indexes = list(self.listbox.curselection())
        if len(indexes) != 1:
            return
        index = indexes[0]
        target = index + step
        if target < 0 or target >= len(self.files):
            return
        self.files[index], self.files[target] = self.files[target], self.files[index]
        self.assignments[index], self.assignments[target] = self.assignments[target], self.assignments[index]
        self._refresh_list()
        self.listbox.selection_set(target)

    def clear_files(self) -> None:
        self.files = []
        self.assignments = []
        self._user_grouped = False
        self._refresh_list()

    def apply_filename_groups(self) -> None:
        self._user_grouped = False
        self._auto_assign()
        self._refresh_list()

    def apply_even_split(self) -> None:
        try:
            n_groups = int(self.n_groups_var.get())
        except ValueError:
            messagebox.showinfo("分组数无效", "均分数必须是整数。")
            return
        if not self.files:
            return
        grouped = partition_even(len(self.files), n_groups)
        self._set_assignments(assignments_from_groups(len(self.files), grouped), user=True)

    def apply_single_sheet(self) -> None:
        if not self.files:
            return
        self._set_assignments(["all"] * len(self.files), user=True)

    def move_selected_to_group(self) -> None:
        name = self.group_choice.get().strip()
        if not name:
            messagebox.showinfo("没有目标组", "先按文件名分组或均分，再把选中项移入某组。")
            return
        indexes = list(self.listbox.curselection())
        if not indexes:
            return
        for index in indexes:
            self.assignments[index] = name
        self._set_assignments(self.assignments, user=True)
        for index in indexes:
            self.listbox.selection_set(index)

    def rename_selected_group(self) -> None:
        new_name = self.rename_var.get().strip()
        if not new_name:
            messagebox.showinfo("组名是空的", "输入新的 sheet 名再改名。")
            return
        indexes = list(self.listbox.curselection())
        if indexes:
            old_names = {self.assignments[index] for index in indexes}
        else:
            current = self.group_choice.get().strip()
            old_names = {current} if current else set()
        if not old_names:
            return
        self.assignments = [new_name if name in old_names else name for name in self.assignments]
        self._set_assignments(self.assignments, user=True)

    def _export_groups(self) -> list[tuple[str, list[Spectrum]]]:
        spectra = self._parsed_spectra()
        grouped_paths = groups_from_assignments(self.files, self.assignments)
        by_path = {item.path.resolve(): item for item in spectra}
        return [
            (name, [by_path[path.resolve()] for path in paths])
            for name, paths in grouped_paths
        ]

    def export(self) -> None:
        if not self.files:
            messagebox.showinfo("没有谱线", "先拖入或添加 txt 文件。")
            return
        layout = self.layout_var.get()
        default_dir = self.files[0].parent
        try:
            resolved = resolve_layout(self._parsed_spectra(), layout)
            default_name = f"origin_{resolved}.opju"
        except ValueError:
            default_name = "origin_spectra.opju"
        output = filedialog.asksaveasfilename(
            title="保存 Origin 工程",
            defaultextension=".opju",
            initialdir=str(default_dir),
            initialfile=default_name,
            filetypes=[("Origin 工程", "*.opju")],
        )
        if not output:
            return
        if origin_process_running():
            ok = messagebox.askokcancel(
                "Origin 正在运行",
                "检测到 Origin64.exe 正在运行。\n"
                "本工具会另开/连接 Origin，新建空工程后写入谱线并保存为 .opju。\n"
                "如果连到了你正在用的 Origin，未保存的工作可能丢失。\n\n继续吗？",
            )
            if not ok:
                return
        self.status.config(text="正在启动 Origin 并写入工作表，请稍等…")
        self.root.update_idletasks()
        try:
            groups = self._export_groups()
            opju_path, xlsx_path, csv_path = export_origin_project(
                self.files,
                Path(output),
                layout=layout,
                groups=groups,
                show_origin=self.show_origin_var.get(),
                keep_open=self.keep_open_var.get(),
                also_xlsx=self.also_xlsx_var.get(),
            )
        except (ValueError, OriginExportError) as exc:
            self._refresh_status()
            messagebox.showerror("Origin 工程生成失败", str(exc))
            return
        self._refresh_status()
        extra = ""
        if xlsx_path:
            extra += f"\nXLSX：{xlsx_path}"
        if csv_path:
            extra += f"\nCSV：{csv_path}"
        messagebox.showinfo("已生成 Origin 工程", f"OPJU：{opju_path}{extra}")

    def run(self) -> None:
        self.root.mainloop()


def _enable_windows_dpi() -> None:
    if sys.platform != "win32":
        return
    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        return


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="把两列 1D 谱线直接导入 Origin，生成 XYYY/XYXY 的 .opju")
    parser.add_argument("--cli", action="store_true", help="不打开窗口，直接导出")
    parser.add_argument("-i", "--input", nargs="+", help="txt 文件或文件夹")
    parser.add_argument("-o", "--output", help="输出 .opju 路径")
    parser.add_argument("--layout", choices=["auto", "XYYY", "XYXY"], default="auto")
    parser.add_argument("--split-temp", action="store_true", help="按文件名自动分组（不限于温度）")
    parser.add_argument("--n-groups", type=int, default=None, help="按列表顺序均分成 N 张 sheet")
    parser.add_argument("--also-xlsx", action="store_true", help="同时写出 xlsx/csv 备份")
    parser.add_argument("--show-origin", action="store_true", help="生成时显示 Origin 窗口")
    parser.add_argument("--keep-open", action="store_true", help="保存后不关闭 Origin")
    parser.add_argument("--xlsx-only", action="store_true", help="只写 xlsx/csv，不启动 Origin")
    return parser


def run_cli(args: argparse.Namespace) -> int:
    if not args.input or not args.output:
        print("CLI 需要 --input 和 --output", file=sys.stderr)
        return 2
    files: list[Path] = []
    for raw in args.input:
        files = load_from_drop_payload(files, [raw])
    output = Path(args.output)
    if origin_process_running() and not args.xlsx_only:
        print("警告：检测到 Origin 正在运行。本工具会新建空工程，未保存的 Origin 工作可能丢失。", file=sys.stderr)
    try:
        if args.xlsx_only:
            xlsx_path, csv_path = export_spectra(
                files,
                output.with_suffix(".xlsx"),
                layout=args.layout,
                split_temp=args.split_temp,
                n_groups=args.n_groups,
                write_csv_copy=True,
            )
            print(xlsx_path)
            if csv_path:
                print(csv_path)
            return 0
        opju_path, xlsx_path, csv_path = export_origin_project(
            files,
            output,
            layout=args.layout,
            split_temp=args.split_temp,
            n_groups=args.n_groups,
            show_origin=args.show_origin,
            keep_open=args.keep_open,
            also_xlsx=args.also_xlsx,
        )
    except (ValueError, OriginExportError) as exc:
        print(f"Origin 工程生成失败：{exc}", file=sys.stderr)
        return 4
    print(opju_path)
    if xlsx_path:
        print(xlsx_path)
    if csv_path:
        print(csv_path)
    return 0


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.cli:
        return run_cli(args)
    _enable_windows_dpi()
    initial: list[Path] = []
    if args.input:
        initial = load_from_drop_payload([], args.input)
    SpectraToOriginApp(initial).run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
